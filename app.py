import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import timedelta
import re, io, base64, os, json

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill

st.set_page_config(page_title="CSU Gradebook + Echo Dashboard", page_icon="🎓", layout="wide")

# ========================= CSU Branding + Persistence =========================
BRAND_FILE = "branding.json"
ASSETS_DIR = "assets"
DEFAULTS = {
    "primary": "#1E4D2B",   # CSU Green
    "gold":    "#C8C372",   # CSU Gold
    "text":    "#111111",
    "bg":      "#FFFFFF",   # White
    "card":    "#F6F7F4",
    "mode":    "light",
    "logo":    f"{ASSETS_DIR}/csu_logo.png"
}

def load_branding():
    if os.path.exists(BRAND_FILE):
        try:
            with open(BRAND_FILE, "r") as f:
                data = json.load(f)
                return {**DEFAULTS, **data}
        except Exception:
            return DEFAULTS.copy()
    return DEFAULTS.copy()

def save_branding(cfg: dict, uploaded_logo=None):
    # Save colors/mode to branding.json
    with open(BRAND_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    # Save uploaded logo (if provided)
    if uploaded_logo is not None:
        logo_path = os.path.join(ASSETS_DIR, "csu_logo.png")
        with open(logo_path, "wb") as out:
            out.write(uploaded_logo.getbuffer())
        cfg["logo"] = logo_path
        with open(BRAND_FILE, "w") as f:
            json.dump(cfg, f, indent=2)
    return cfg

def load_logo_b64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        return None

def inject_brand_css(primary, gold, text, bg, card, mode, logo_b64=None):
    if mode == "dark":
        bg="#0E1117"; card="#161B22"; text="#E6EDF3"
    logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:42px;margin-right:10px;vertical-align:middle;" />' if logo_b64 else ""
    st.markdown(f"""
    <style>
    :root {{
      --brand-primary: {primary};
      --brand-gold: {gold};
      --brand-text: {text};
      --brand-bg: {bg};
      --brand-card: {card};
    }}
    .stApp {{ background: var(--brand-bg); color: var(--brand-text); }}
    .kpi-card {{
      background: var(--brand-card); border-radius: 14px; padding: 14px;
      box-shadow: 0 4px 18px rgba(0,0,0,0.06); border: 1px solid rgba(0,0,0,0.05);
    }}
    .kpi-title {{font-size: 0.9rem; color: #4b5563; margin-bottom: 4px;}}
    .kpi-value {{font-size: 1.6rem; font-weight: 700; color: var(--brand-text);}}
    .stTabs [data-baseweb="tab-list"] {{gap: 8px;}}
    .stTabs [data-baseweb="tab"] {{padding: 10px 16px; background: #ffffff; border-radius: 10px;
                                  border: 1px solid rgba(0,0,0,0.08);}}
    .stTabs [aria-selected="true"] {{background: #eef5ee; border-color: var(--brand-primary);}}
    .gold-badge {{ color: var(--brand-primary); background: rgba(200,195,114,0.15); padding:2px 8px; border-radius: 999px; font-weight:600; }}
    </style>
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
      {logo_html}
      <span style="font-weight:700;font-size:1.25rem;color:var(--brand-text)">
        Colorado State — Gradebook & Echo Dashboard
      </span>
      <span class="gold-badge">CSU</span>
    </div>
    """, unsafe_allow_html=True)

def apply_colorway(fig):
    fig.update_layout(colorway=[st.session_state.brand["primary"], st.session_state.brand["gold"], "#2E7D32", "#8C7E3E", "#0B5D1E"])
    return fig

# Load or init branding
if "brand" not in st.session_state:
    st.session_state.brand = load_branding()

# Sidebar branding controls
with st.sidebar:
    st.markdown("### Theme & Branding")
    mode = st.radio("Theme", ["Light","Dark"], horizontal=True, index=0 if st.session_state.brand["mode"]=="light" else 1)
    primary = st.color_picker("Primary (CSU Green)", st.session_state.brand["primary"])
    gold    = st.color_picker("Accent (CSU Gold)", st.session_state.brand["gold"])
    text    = st.color_picker("Text", st.session_state.brand["text"])
    bg      = st.color_picker("Background", st.session_state.brand["bg"])
    card    = st.color_picker("Card Background", st.session_state.brand["card"])
    logo_upload = st.file_uploader("Logo (PNG)", type=["png"])

    if st.button("💾 Save branding"):
        st.session_state.brand.update({"primary":primary, "gold":gold, "text":text, "bg":bg, "card":card, "mode": mode.lower()})
        st.session_state.brand = save_branding(st.session_state.brand, uploaded_logo=logo_upload)
        st.success("Branding saved. It will auto-load on next run.")

# Apply CSS based on current brand
logo_b64 = load_logo_b64(st.session_state.brand["logo"]) if st.session_state.brand.get("logo") else None
inject_brand_css(primary=primary, gold=gold, text=text, bg=bg, card=card, mode=mode.lower(), logo_b64=logo_b64)

# ========================= Shared helpers =========================
def kpi_card(title, value):
    st.markdown(f"""<div class="kpi-card"><div class="kpi-title">{title}</div><div class="kpi-value">{value}</div></div>""", unsafe_allow_html=True)

def fig_layout(fig, h=420):
    fig.update_layout(height=h, margin=dict(l=30,r=10,t=40,b=35),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    return apply_colorway(fig)

def shorten_label(s, limit=40):
    s2 = re.sub(r"\(\d+\)$", "", str(s)).strip()
    return (s2[:limit] + "…") if len(s2) > limit else s2

# ========================= Echo Script Parity =========================
def time_to_seconds(ts: str) -> int:
    if pd.isna(ts) or ts == "":
        return 0
    parts = list(map(int, str(ts).split(":")))
    while len(parts) < 3:
        parts.insert(0, 0)
    h, m, s = parts
    return h*3600 + m*60 + s

def seconds_to_hms(sec: float) -> str:
    return "" if pd.isna(sec) else str(timedelta(seconds=int(sec)))

def natural_key(s: str):
    return [int(chunk) if chunk.isdigit() else chunk.lower() for chunk in re.split(r'(\d+)', str(s))]

def echo_analyze(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df['Duration_sec']      = df['Duration'].apply(time_to_seconds)
    df['TotalViewTime_sec'] = df['Total View Time'].apply(time_to_seconds)
    df['AvgViewTime_sec']   = df['Average View Time'].apply(time_to_seconds)
    df['Row View %']        = df['TotalViewTime_sec'] / df['Duration_sec'].replace(0, np.nan)

    grp = df.groupby('Media Name', sort=False)
    titles = list(grp.groups.keys())

    summary_core = pd.DataFrame({
        'Media Title':              titles,
        'Video Duration':           [grp.get_group(t)['Duration_sec'].iloc[0] for t in titles],
        'Number of Unique Viewers': grp['User Name'].nunique().values,
        'Average View %':           grp['Row View %'].mean().fillna(0).values,
        'Total View %':             (grp['TotalViewTime_sec'].sum() / grp['Duration_sec'].sum()).values,
        'Total View Time':          grp['TotalViewTime_sec'].sum().values,
        'Average View Time':        grp['AvgViewTime_sec'].mean().values,
        'Average Total View Time':  grp['TotalViewTime_sec'].mean().values,
    })

    summary_core['sort_key'] = summary_core['Media Title'].apply(natural_key)
    summary_core = summary_core.sort_values('sort_key').drop(columns='sort_key').reset_index(drop=True)

    means = summary_core[['Video Duration','Total View Time','Average View Time','Average Total View Time']].mean()
    viewers_mean = summary_core['Number of Unique Viewers'].mean()
    summary_core.loc[len(summary_core)] = {
        'Media Title':               'Grand Total',
        'Video Duration':            means['Video Duration'],
        'Number of Unique Viewers':  viewers_mean,
        'Average View %':            summary_core['Average View %'].mean(),
        'Total View %':              summary_core['Total View %'].mean(),
        'Total View Time':           means['Total View Time'],
        'Average View Time':         means['Average View Time'],
        'Average Total View Time':   means['Average Total View Time'],
    }

    n = len(summary_core) - 1
    means2 = summary_core.loc[:n-1, ['Video Duration','Total View Time','Average View Time','Average Total View Time']].mean()
    summary_core.loc[len(summary_core)] = {
        'Media Title':               'Average Video Length and Watch Time',
        'Video Duration':            means2['Video Duration'],
        'Number of Unique Viewers':  '',
        'Average View %':            summary_core.loc[:n-1, 'Average View %'].mean(),
        'Total View %':              summary_core.loc[:n-1, 'Total View %'].mean(),
        'Total View Time':           means2['Total View Time'],
        'Average View Time':         means2['Average View Time'],
        'Average Total View Time':   means2['Average Total View Time'],
    }
    return summary_core

def echo_build_workbook(summary_df: pd.DataFrame) -> bytes:
    # NOTE: Conditional formatting colors remain Excel defaults for accessibility.
    wb = Workbook()
    ws = wb.active
    ws.title = 'Media Summary'

    tmp = summary_df.copy()
    for col in ['Video Duration','Total View Time','Average View Time','Average Total View Time']:
        tmp[col] = tmp[col].apply(seconds_to_hms)
    for row in dataframe_to_rows(tmp, index=False, header=True):
        ws.append(row)
    last_row = ws.max_row
    media_count = len(summary_df) - 2

    for r in range(2, last_row + 1):
        cell = ws[f'B{r}']
        secs = time_to_seconds(cell.value)
        cell.value = secs / 86400.0
        cell.number_format = 'hh:mm:ss'

    for r in range(2, last_row + 1):
        for col in ('D','E'):
            c = ws[f'{col}{r}']
            if isinstance(c.value, (int, float)):
                c.number_format = '0.00%'
        for col in ('F','G','H'):
            ws[f'{col}{r}'].number_format = 'hh:mm:ss'

    if media_count >= 1:
        bar = DataBarRule(start_type='min', end_type='max')  # default color
        ws.conditional_formatting.add(f"B2:B{1+media_count}", bar)
        ws.conditional_formatting.add(f"D2:D{1+media_count}", bar)

    chart1 = LineChart()
    chart1.title = "View % Over Time"
    chart1.style = 9
    chart1.y_axis.number_format = '0.00%'
    data1 = Reference(ws, min_col=4, min_row=1, max_row=1+media_count)
    chart1.add_data(data1, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=2, max_row=1+media_count)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "J2")

    chart2 = LineChart()
    chart2.title = "Unique Viewers Over Time"
    chart2.style = 9
    data2 = Reference(ws, min_col=3, min_row=1, max_row=1+media_count)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "J20")

    tbl = Table(displayName="MediaStats", ref=f"A1:H{last_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ========================= Gradebook Script Parity =========================
def gradebook_process(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    mask = df.iloc[:, 0].astype(str).str.contains("Student, Test", na=False)
    df = df[~mask].reset_index(drop=True)
    to_drop = ["Student","ID","SIS User ID","SIS Login ID","Current Grade","Unposted Current Grade","Unposted Final Grade"]
    df.drop(columns=[c for c in to_drop if c in df.columns], inplace=True, errors="ignore")

    drop_cols = []
    for col in df.columns:
        if col == "Final Grade":
            continue
        s = pd.to_numeric(df[col].iloc[2:], errors='coerce')
        if s.fillna(0).eq(0).all():
            drop_cols.append(col)
    df.drop(columns=drop_cols, inplace=True, errors="ignore")
    return df

def gradebook_build_workbook(df: pd.DataFrame) -> bytes:
    # NOTE: Conditional formatting colors remain Excel defaults for accessibility.
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    final_grade_idx_pre = None
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == "Final Grade":
            final_grade_idx_pre = ci
            break

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.column == final_grade_idx_pre:
                continue
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column == final_grade_idx_pre:
                continue
            if isinstance(cell.value, str) and cell.value.strip():
                txt = cell.value.replace(",", "")
                try:
                    cell.value = float(txt)
                except ValueError:
                    pass

    data_last_row = ws.max_row
    for ci in range(1, ws.max_column + 1):
        if ci == final_grade_idx_pre:
            continue
        hdr = ws.cell(row=2, column=ci)
        if isinstance(hdr.value, str) and "(read only)" in hdr.value:
            nums = [ws.cell(row=r, column=ci).value for r in range(3, data_last_row + 1) if isinstance(ws.cell(row=r, column=ci).value, (int, float))]
            if nums:
                hdr.value = max(nums)

    ws.insert_cols(1)
    ws["A1"] = "Row Titles"

    final_grade_idx = None
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == "Final Grade":
            final_grade_idx = ci
            break

    ws["A2"] = "Points Possible"
    original_last_data_row = ws.max_row
    avg_row  = original_last_data_row + 1
    avg0_row = original_last_data_row + 2
    ws[f"A{avg_row}"]  = "Average"
    ws[f"A{avg0_row}"] = "Average Excluding Zeros"

    max_col = ws.max_column
    for col in range(2, max_col + 1):
        if col == final_grade_idx:
            continue
        letter   = get_column_letter(col)
        data_rng = f"{letter}3:{letter}{original_last_data_row}"
        header   = f"{letter}$2"
        c_avg    = ws[f"{letter}{avg_row}"]
        c_avg.value = f"=AVERAGE({data_rng})/{header}"
        c_avg.number_format = '0.00%'
        c_avg0   = ws[f"{letter}{avg0_row}"]
        c_avg0.value = f"=AVERAGEIF({data_rng},\">0\")/{header}"
        c_avg0.number_format = '0.00%'

    green  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    for row in (avg_row, avg0_row):
        rng = f"B{row}:{get_column_letter(max_col)}{row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0.9'], fill=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator='between',     formula=['0.8','0.9'], fill=yellow))
        ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan',    formula=['0.8'], fill=red))

    count_row = avg0_row + 1
    pct_row   = avg0_row + 2
    ws[f"A{count_row}"] = "Count of F"
    ws[f"A{pct_row}"]   = "Percent of F"
    fg_letter = get_column_letter(final_grade_idx)
    ws.cell(row=count_row, column=final_grade_idx).value = f'=COUNTIF({fg_letter}3:{fg_letter}{original_last_data_row},"F")'
    total_students = original_last_data_row - 2
    ws.cell(row=pct_row, column=final_grade_idx).value = f'={fg_letter}{count_row}/{total_students}'
    ws.cell(row=pct_row, column=final_grade_idx).number_format = '0.00%'

    table_end = get_column_letter(max_col) + str(original_last_data_row)
    table = Table(displayName="GradesTable", ref=f"A1:{table_end}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ========================= Gradebook Dashboard Prep =========================
def parse_missing_excused(series: pd.Series):
    s = series.astype(str).str.strip()
    excused = s.str.upper().isin(["EX","EXCUSED"])
    blanks  = s.isin(["","nan","NaN","-","—","–"])
    nums = pd.to_numeric(s.str.replace("%","", regex=False), errors="coerce")
    missing = blanks | nums.isna()
    return nums, missing, excused

def extract_category(c):
    base = re.sub(r"\(\d+\)$", "", c).strip()
    if ":" in base: return base.split(":")[0].strip()
    if "-" in base: return base.split("-")[0].strip()
    return base.split()[0] if base.split() else base

def gradebook_prepare_for_dashboard(df_proc: pd.DataFrame, student_col="Student", section_col="Section", final_col="Final Grade"):
    if df_proc.shape[0] < 3:
        return None
    points = df_proc.iloc[1].copy()
    data   = df_proc.iloc[2:].copy()
    if student_col in df_proc.columns:
        data[student_col] = df_proc.iloc[2:][student_col].values
    if section_col in df_proc.columns:
        data[section_col] = df_proc.iloc[2:][section_col].values

    assign_cols = [c for c in df_proc.columns if c not in [student_col, section_col, final_col]]
    num_df = pd.DataFrame(index=data.index, columns=assign_cols, dtype=float)
    miss_df = pd.DataFrame(False, index=data.index, columns=assign_cols)
    exc_df = pd.DataFrame(False, index=data.index, columns=assign_cols)
    for c in assign_cols:
        nums, miss, exc = parse_missing_excused(data[c])
        num_df[c] = nums
        miss_df[c] = miss | nums.isna()
        exc_df[c] = exc

    pts = pd.to_numeric(points[assign_cols], errors="coerce").replace(0, np.nan)
    pct_df = (num_df / pts) * 100.0
    return {
        "data": data.reset_index(drop=True),
        "final_col": final_col if final_col in df_proc.columns else None,
        "student_col": student_col if student_col in df_proc.columns else None,
        "section_col": section_col if section_col in df_proc.columns else None,
        "assign_cols": assign_cols,
        "pct_df": pct_df,
        "missing": miss_df,
        "excused": exc_df,
        "points": pts,
        "categories": {c: extract_category(c) for c in assign_cols}
    }

# ========================= Main Layout =========================
tab_gb, tab_echo, tab_profile = st.tabs(["📘 Gradebook", "🎬 Echo", "👤 Student Profile"])

# ===== Gradebook Tab =====
with tab_gb:
    st.subheader("Upload a Gradebook CSV")
    gb = st.file_uploader("Choose a gradebook CSV", type=["csv"], key="gb_up_csu_persist")

    with st.expander("Column mapping (Gradebook)"):
        student_col = st.text_input("Student column", value="Student")
        section_col = st.text_input("Section column (optional)", value="Section")
        final_col   = st.text_input("Final letter grade column", value="Final Grade")

    if not gb:
        st.info("Upload a gradebook CSV (Canvas export). Ensure row 2 is Points Possible.")
    else:
        df_raw = pd.read_csv(gb)
        rename_map = {}
        if student_col in df_raw.columns: rename_map[student_col] = "Student"
        if section_col in df_raw.columns and section_col != "": rename_map[section_col] = "Section"
        if final_col in df_raw.columns: rename_map[final_col] = "Final Grade"
        df_std = df_raw.rename(columns=rename_map)

        df_proc = gradebook_process(df_std)
        st.write("**Processed preview**")
        st.dataframe(df_proc.head(20), use_container_width=True)

        try:
            xbytes = gradebook_build_workbook(df_proc)
            st.download_button("⬇️ Download Excel (script formulas & formatting)",
                               data=xbytes, file_name="Gradebook_Analyzed.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Excel build failed: {e}")

        st.markdown("---")
        st.subheader("Interactive Dashboard Views")

        prep = gradebook_prepare_for_dashboard(df_proc, student_col="Student", section_col="Section", final_col="Final Grade")
        if not prep:
            st.info("Not enough rows to compute dashboard metrics.")
        else:
            data = prep["data"]
            final_c = prep["final_col"]
            student_c = prep["student_col"]
            section_c = prep["section_col"]
            assign_cols = prep["assign_cols"]
            pct_df = prep["pct_df"]
            missing = prep["missing"]
            excused = prep["excused"]
            categories = prep["categories"]

            if "gb_filters" not in st.session_state:
                st.session_state.gb_filters = {"section":"All", "students":[], "categories":[]}

            fc1, fc2, fc3, fc4 = st.columns([1.2,1.2,2,1])
            with fc1:
                if section_c and section_c in data.columns:
                    opts = ["All"] + sorted([x for x in data[section_c].dropna().unique().tolist() if x!=""])
                    sel = st.session_state.gb_filters.get("section","All")
                    idx = opts.index(sel) if sel in opts else 0
                    st.session_state.gb_filters["section"] = st.selectbox("Section", opts, index=idx)
            with fc2:
                cat_list = sorted(set(categories.values()))
                st.session_state.gb_filters["categories"] = st.multiselect("Categories", options=cat_list, default=st.session_state.gb_filters["categories"])
            with fc3:
                if student_c and student_c in data.columns:
                    options = data[student_c].dropna().astype(str).tolist()
                    prev = st.session_state.gb_filters.get("students", [])
                    default = [s for s in prev if s in options][:8]
                    picks = st.multiselect("Students (focus)", options=options, default=default, max_selections=8)
                    st.session_state.gb_filters["students"] = picks
            with fc4:
                st.caption("Filters persist within this tab")

            rows_mask = pd.Series(True, index=data.index)
            if st.session_state.gb_filters["section"] != "All" and section_c and section_c in data.columns:
                rows_mask &= (data[section_c] == st.session_state.gb_filters["section"])
            data_f = data.loc[rows_mask].reset_index(drop=True)
            pct_f = pct_df.loc[rows_mask]
            miss_f = missing.loc[rows_mask]
            exc_f  = excused.loc[rows_mask]

            assn_f = assign_cols[:]
            if st.session_state.gb_filters["categories"]:
                assn_f = [a for a in assign_cols if categories[a] in st.session_state.gb_filters["categories"]]
            if len(assn_f) == 0:
                assn_f = assign_cols[:]

            c1, c2, c3, c4 = st.columns(4)
            kpi_card("Students", data_f.shape[0])
            kpi_card("Assignments", len(assn_f))
            if final_c and final_c in data_f.columns:
                vals = data_f[final_c].astype(str)
                kpi_card("Unique Letter Grades", vals.nunique())
                kpi_card("Count of F", int((vals == "F").sum()))
            else:
                kpi_card("Unique Letter Grades", "—")
                kpi_card("Count of F", "—")

            incl = pct_f[assn_f].fillna(0.0)
            excl = pct_f[assn_f].copy()
            avg_incl = incl.mean().sort_values()
            avg_excl = excl.mean().reindex(avg_incl.index)
            labels = [shorten_label(a, 36) for a in avg_incl.index]
            fig = go.Figure()
            fig.add_bar(x=labels, y=avg_incl.values, name="Including Missing")
            fig.add_bar(x=labels, y=avg_excl.values, name="Excluding Missing")
            fig.update_xaxes(tickangle=45)
            st.plotly_chart(fig_layout(fig, h=420), use_container_width=True)

            st.subheader("Missing / Excused Heatmap")
            if not miss_f.empty and student_c and student_c in data_f.columns:
                mat = miss_f[assn_f].astype(int).values.astype(float) - 0.5 * exc_f[assn_f].astype(int).values
                fig_hm = px.imshow(
                    mat,
                    labels=dict(x="Assignments", y="Students", color="Status"),
                    x=[shorten_label(a, 18) for a in assn_f],
                    y=data_f[student_c].tolist(),
                    aspect="auto"
                )
                st.plotly_chart(fig_layout(fig_hm, h=min(800, 40 + 22*data_f.shape[0])), use_container_width=True)

            st.subheader("Student Trajectories")
            if student_c and student_c in data_f.columns:
                picks = st.session_state.gb_filters["students"]
                if picks:
                    figT = go.Figure()
                    for s in picks:
                        row = pct_f.loc[data_f[student_c]==s, assn_f]
                        if not row.empty:
                            figT.add_scatter(x=[shorten_label(a, 36) for a in assn_f], y=row.iloc[0].values, mode="lines+markers", name=s)
                    figT.update_yaxes(title_text="Score (%)", range=[0,100])
                    figT.update_xaxes(tickangle=45, title_text="Assignments")
                    st.plotly_chart(fig_layout(figT, h=420), use_container_width=True)
                else:
                    st.caption("Select students above to view trajectories.")

            st.subheader("Assignment Correlations (completed work only)")
            if len(assn_f) >= 2:
                corr = pct_f[assn_f].replace(0, np.nan).corr()
                st.plotly_chart(fig_layout(px.imshow(corr, text_auto=".2f", aspect="auto", color_continuous_scale="Greens"), h=520), use_container_width=True)

            st.subheader("Weights & Scenarios")
            cat_list = sorted(set(categories.values()))
            if cat_list and not pct_f.empty:
                cols = st.columns(min(4, len(cat_list)))
                weights = {}
                for i, cat in enumerate(cat_list):
                    with cols[i % len(cols)]:
                        weights[cat] = st.slider(f"{cat}", 0.0, 100.0, 100.0/len(cat_list), 1.0)
                total = sum(weights.values()) if weights else 0.0
                st.write(f"**Total = {total:.1f}%** (normalized)")

                if total > 0:
                    wnorm = {k: v/total for k, v in weights.items()}
                    per_cat = {}
                    for cat in cat_list:
                        cols_cat = [a for a in assn_f if categories[a] == cat]
                        if cols_cat:
                            per_cat[cat] = pct_f[cols_cat].replace(0, np.nan).mean(axis=1)
                    wfinal = None
                    for cat, series in per_cat.items():
                        contrib = series * wnorm.get(cat, 0.0)
                        wfinal = contrib if wfinal is None else (wfinal + contrib)
                    comp = pd.DataFrame({
                        student_c if student_c else "Index": data_f[student_c] if student_c else np.arange(len(pct_f)),
                        "What-if Final (%)": wfinal.round(1) if wfinal is not None else np.nan
                    })
                    st.dataframe(comp, use_container_width=True)

                if "gb_scenarios" not in st.session_state: st.session_state.gb_scenarios = {}
                name = st.text_input("Scenario name", value="CSU Scenario")
                if st.button("Save scenario"):
                    st.session_state.gb_scenarios[name] = weights.copy()
                    st.success(f"Saved scenario '{{name}}'")

                if st.session_state.gb_scenarios:
                    st.subheader("Saved Scenarios")
                    st.json(st.session_state.gb_scenarios)

# ===== Echo Tab =====
with tab_echo:
    st.subheader("Upload an Echo CSV")
    ec = st.file_uploader("Choose an Echo CSV", type=["csv"], key="echo_up_csu_persist")

    with st.expander("Column mapping (Echo)"):
        c_media = st.text_input("Media Title column", value="Media Name")
        c_dur   = st.text_input("Duration column", value="Duration")
        c_user  = st.text_input("User column", value="User Name")
        c_tot   = st.text_input("Total View Time column", value="Total View Time")
        c_avg   = st.text_input("Average View Time column", value="Average View Time")

    if not ec:
        st.info("Upload an Echo CSV (Panopto/Echo360 export).")
    else:
        df = pd.read_csv(ec, dtype=str)
        rename_map = {}
        if c_media in df.columns: rename_map[c_media] = "Media Name"
        if c_dur in df.columns:   rename_map[c_dur]   = "Duration"
        if c_user in df.columns:  rename_map[c_user]  = "User Name"
        if c_tot in df.columns:   rename_map[c_tot]   = "Total View Time"
        if c_avg in df.columns:   rename_map[c_avg]   = "Average View Time"
        df = df.rename(columns=rename_map)

        missing_cols = [c for c in ["Media Name","Duration","User Name","Total View Time","Average View Time"] if c not in df.columns]
        if missing_cols:
            st.error(f"Missing required columns: {missing_cols}")
        else:
            summary = echo_analyze(df)

            try:
                xbytes = echo_build_workbook(summary)
                st.download_button("⬇️ Download Excel (script formatting & charts)",
                                   data=xbytes, file_name="Echo_Analyzed.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Excel build failed: {e}")

            st.markdown("---")
            st.subheader("Interactive Dashboard Views")
            disp = summary.copy()
            for col in ['Video Duration','Total View Time','Average View Time','Average Total View Time']:
                disp[col] = disp[col].apply(lambda x: seconds_to_hms(x) if isinstance(x, (int,float,np.integer,np.floating)) else x)
            st.dataframe(disp, use_container_width=True)

            media_count = max(0, len(summary) - 2)
            if media_count > 0:
                main = summary.iloc[:media_count].copy()
                fig1 = go.Figure()
                fig1.add_trace(go.Scatter(x=main['Media Title'], y=main['Average View %'], mode='lines+markers', name='Average View %'))
                fig1.update_yaxes(tickformat=".0%")
                st.plotly_chart(fig_layout(fig1, h=420), use_container_width=True)

                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(x=main['Media Title'], y=main['Number of Unique Viewers'], mode='lines+markers', name='Unique Viewers'))
                st.plotly_chart(fig_layout(fig2, h=420), use_container_width=True)

# ===== Student Profile Tab =====
with tab_profile:
    st.subheader("Student Profile (from current Gradebook upload)")
    gb2 = st.file_uploader("Re-select Gradebook CSV (for this tab)", type=["csv"], key="gb_profile_csu_persist")
    if gb2:
        df_raw2 = pd.read_csv(gb2)
        student_col_p = st.text_input("Student column (profile)", value="Student", key="student_col_p")
        section_col_p = st.text_input("Section column (profile)", value="Section", key="section_col_p")
        final_col_p   = st.text_input("Final letter grade column (profile)", value="Final Grade", key="final_col_p")

        rename_map = {}
        if student_col_p in df_raw2.columns: rename_map[student_col_p] = "Student"
        if section_col_p in df_raw2.columns and section_col_p != "": rename_map[section_col_p] = "Section"
        if final_col_p in df_raw2.columns: rename_map[final_col_p] = "Final Grade"
        df_std2 = df_raw2.rename(columns=rename_map)
        df_proc2 = gradebook_process(df_std2)
        prep2 = gradebook_prepare_for_dashboard(df_proc2, student_col="Student", section_col="Section", final_col="Final Grade")
        if not prep2:
            st.info("Not enough rows to compute student profile.")
        else:
            data = prep2["data"]; pct_df = prep2["pct_df"]; missing = prep2["missing"]; excused = prep2["excused"]; assign_cols = prep2["assign_cols"]
            students = data["Student"].dropna().astype(str).tolist() if "Student" in data.columns else []
            who = st.selectbox("Choose a student", students)
            if who:
                row = data.loc[data["Student"]==who].reset_index(drop=True)
                pct_row = pct_df.loc[data["Student"]==who, assign_cols]
                if not pct_row.empty:
                    fig = go.Figure()
                    fig.add_bar(x=[shorten_label(a, 36) for a in assign_cols], y=pct_row.iloc[0].values, name="Score %")
                    fig.update_yaxes(range=[0,100]); fig.update_xaxes(tickangle=45)
                    st.plotly_chart(fig_layout(fig, h=420), use_container_width=True)

                    status = pd.DataFrame({
                        "Assignment": assign_cols,
                        "Missing": missing.loc[data["Student"]==who, assign_cols].iloc[0].values,
                        "Excused": excused.loc[data["Student"]==who, assign_cols].iloc[0].values,
                    })
                    status["Status"] = np.where(status["Excused"], "Excused", np.where(status["Missing"], "Missing", "Submitted"))
                    st.dataframe(status[["Assignment","Status"]], use_container_width=True)

                    miss_count = (status["Status"]=="Missing").sum()
                    exc_count  = (status["Status"]=="Excused").sum()
                    kpi_card("Assignments", len(assign_cols))
                    kpi_card("Missing", miss_count)
                    kpi_card("Excused", exc_count)
    else:
        st.info("Upload a gradebook in this tab to view a student profile.")